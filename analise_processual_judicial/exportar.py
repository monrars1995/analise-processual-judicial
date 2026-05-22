#!/usr/bin/env python3
"""
Exporta dados do processo judicial para XLSX e CSV.
Útil para análise em planilhas e integração com outros sistemas.
"""
import json
import sys
import csv
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def exportar_csv(dados, saida_csv):
    linha = dados.get('linha_do_tempo', [])
    with open(saida_csv, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(['Data', 'Hora', 'Tipo', 'Documento', 'Origem'])
        for item in linha:
            writer.writerow([
                item.get('data', ''),
                item.get('hora', ''),
                item.get('tipo', ''),
                item.get('documento', ''),
                item.get('origem', '')
            ])
    print(f"CSV exportado: {saida_csv}")


def exportar_xlsx(dados, irregularidades, saida_xlsx):
    if not HAS_OPENPYXL:
        print("Erro: openpyxl não instalado. Execute: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def estilizar_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # Aba 1: Metadados
    ws_meta = wb.active
    ws_meta.title = "Metadados"
    meta = dados.get('metadados', {})
    campos = [
        ("Número do Processo", meta.get('numero_processo', '')),
        ("Classe", meta.get('classe', '')),
        ("Assunto", meta.get('assunto', '')),
        ("Prioridade", meta.get('prioridade', '')),
        ("Data da Distribuição", meta.get('data_distribuicao', '')),
        ("Valor da Causa", meta.get('valor_causa', '')),
        ("Segredo de Justiça", meta.get('segredo_justica', '')),
    ]
    ws_meta.append(["Campo", "Valor"])
    for rotulo, valor in campos:
        ws_meta.append([rotulo, valor])
    # Partes
    ws_meta.append([])
    ws_meta.append(["Partes"])
    ws_meta.append(["Polo", "Nome"])
    for parte in meta.get('partes', []):
        ws_meta.append([parte.get('polo', ''), parte.get('nome', '')])
    estilizar_header(ws_meta, 1)
    estilizar_header(ws_meta, 9)
    ws_meta.column_dimensions['A'].width = 25
    ws_meta.column_dimensions['B'].width = 50

    # Aba 2: Linha do Tempo
    ws_tl = wb.create_sheet("Linha do Tempo")
    ws_tl.append(["Data", "Hora", "Tipo", "Documento", "Origem"])
    for item in dados.get('linha_do_tempo', []):
        ws_tl.append([
            item.get('data', ''),
            item.get('hora', ''),
            item.get('tipo', ''),
            item.get('documento', ''),
            item.get('origem', '')
        ])
    estilizar_header(ws_tl)
    ws_tl.column_dimensions['A'].width = 12
    ws_tl.column_dimensions['B'].width = 8
    ws_tl.column_dimensions['C'].width = 20
    ws_tl.column_dimensions['D'].width = 50
    ws_tl.column_dimensions['E'].width = 15

    # Aba 3: Irregularidades
    ws_irreg = wb.create_sheet("Irregularidades")
    ws_irreg.append(["#", "Gravidade", "Categoria", "Descrição", "Fundamento", "Data", "Recomendação"])
    for idx, ir in enumerate(irregularidades.get('irregularidades', []), 1):
        cor = None
        if ir['gravidade'] == 'ALTA':
            cor = PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid")
        elif ir['gravidade'] == 'MEDIA':
            cor = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")
        row = [
            idx,
            ir['gravidade'],
            ir['categoria'],
            ir['descricao'],
            ir['fundamento'],
            ir.get('data_ocorrencia') or '',
            ir['recomendacao']
        ]
        ws_irreg.append(row)
        if cor:
            for cell in ws_irreg[ws_irreg.max_row]:
                cell.fill = cor
                cell.border = thin_border
    estilizar_header(ws_irreg)
    ws_irreg.column_dimensions['A'].width = 5
    ws_irreg.column_dimensions['B'].width = 10
    ws_irreg.column_dimensions['C'].width = 20
    ws_irreg.column_dimensions['D'].width = 50
    ws_irreg.column_dimensions['E'].width = 30
    ws_irreg.column_dimensions['F'].width = 12
    ws_irreg.column_dimensions['G'].width = 40

    wb.save(saida_xlsx)
    print(f"XLSX exportado: {saida_xlsx}")


def main():
    if len(sys.argv) < 3:
        print("Uso: python exportar.py <dados.json> <irregularidades.json> [saida.xlsx] [saida.csv]")
        sys.exit(1)

    dados_path = Path(sys.argv[1])
    irreg_path = Path(sys.argv[2])
    saida_xlsx = Path(sys.argv[3]) if len(sys.argv) > 3 else dados_path.with_suffix('.xlsx')
    saida_csv = Path(sys.argv[4]) if len(sys.argv) > 4 else dados_path.with_suffix('.csv')

    with open(dados_path, 'r', encoding='utf-8') as f:
        dados = json.load(f)
    with open(irreg_path, 'r', encoding='utf-8') as f:
        irregularidades = json.load(f)

    exportar_csv(dados, saida_csv)
    exportar_xlsx(dados, irregularidades, saida_xlsx)


if __name__ == '__main__':
    main()
